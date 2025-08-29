import os
import json
from datetime import datetime
from io import BytesIO

from django.conf import settings
from django.utils import timezone

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.platypus import PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from analysis.models import Analysis, Vulnerabilidades
from certs.models import Certificate


class ReportGenerator:
    """
    Generador base de reportes según especificaciones Proyecto Sócrates
    """
    
    def __init__(self, client, filters=None):
        self.client = client
        self.filters = filters or {}
        self.created_at = timezone.now()
    
    def get_certificates_queryset(self):
        """Obtener certificados filtrados del cliente"""
        queryset = Certificate.objects.filter(
            cliente=self.client,
            active=True
        )
        
        # Aplicar filtros adicionales según requerimientos
        if self.filters.get('protocolo'):
            queryset = queryset.filter(protocolo=self.filters['protocolo'])
        
        if self.filters.get('expires_before'):
            queryset = queryset.filter(
                analyses__parametros_generales__fecha_fin__lte=self.filters['expires_before']
            )
        
        return queryset.distinct()
    
    def get_analyses_queryset(self):
        """Obtener análisis exitosos con datos completos"""
        certificates = self.get_certificates_queryset()
        queryset = Analysis.objects.filter(
            certificado__in=certificates,
            tuvo_exito=True
        ).select_related(
            'certificado__cliente',
            'parametros_generales', 
            'parametros_tls',
            'parametros_ssh',
            'parametros_web',
            'otros_parametros',
            'cadena_certificacion'
        ).prefetch_related('vulnerabilidades')
        
        # Filtro por fecha según especificaciones
        if self.filters.get('date_from'):
            queryset = queryset.filter(fecha_inicio__gte=self.filters['date_from'])
        
        if self.filters.get('date_to'):
            queryset = queryset.filter(fecha_inicio__lte=self.filters['date_to'])
        
        return queryset.order_by('-fecha_inicio')


class PDFReportGenerator(ReportGenerator):
    """
    Generador de reportes PDF según especificaciones Proyecto Sócrates
    Incluye TODOS los parámetros definidos en el documento oficial
    """
    
    def generate_certificate_detailed_report(self, file_path):
        """Reporte detallado con TODOS los parámetros especificados"""
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Título principal según documento
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        story.append(Paragraph('Proyecto Sócrates', title_style))
        story.append(Paragraph('Vigilancia activa sobre Certificados Digitales', styles['Heading2']))
        story.append(Paragraph(f'Cliente: {self.client.name}', styles['Heading2']))
        story.append(Paragraph(f'Generado: {self.created_at.strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Resumen ejecutivo
        certificates = self.get_certificates_queryset()
        analyses = self.get_analyses_queryset()
        vulnerabilities = Vulnerabilidades.objects.filter(analisis__in=analyses)
        
        summary_data = [
            ['Métrica', 'Valor'],
            ['Total Certificados', str(certificates.count())],
            ['Análisis Realizados', str(analyses.count())],
            ['Certificados Activos', str(certificates.filter(analyses__otros_parametros__disponibilidad=True).distinct().count())],
            ['Vulnerabilidades Críticas', str(vulnerabilities.filter(severity='CRITICAL').count())],
            ['Vulnerabilidades Altas', str(vulnerabilities.filter(severity='HIGH').count())],
            ['Certificados próximos a expirar (30 días)', str(self._get_expiring_certificates_count(30))],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(self._get_table_style())
        
        story.append(Paragraph('Resumen Ejecutivo', styles['Heading2']))
        story.append(summary_table)
        story.append(PageBreak())
        
        # Detalle por certificado según especificaciones
        for cert in certificates[:10]:  # Limitar para ejemplo
            last_analysis = cert.analyses.filter(tuvo_exito=True).first()
            if last_analysis:
                story.extend(self._generate_certificate_detail(cert, last_analysis, styles))
                story.append(PageBreak())
        
        doc.build(story)
        return file_path
    
    def _generate_certificate_detail(self, certificate, analysis, styles):
        """Generar detalle completo de un certificado según especificaciones"""
        story = []
        target = certificate.ip or certificate.url
        
        story.append(Paragraph(f'Certificado: {target}:{certificate.puerto}', styles['Heading1']))
        story.append(Spacer(1, 12))
        
        # A) Parámetros Generales del Certificado
        if hasattr(analysis, 'parametros_generales') and analysis.parametros_generales:
            params = analysis.parametros_generales
            
            general_data = [
                ['Parámetro', 'Valor'],
                ['Common Name (CN)', params.common_name or 'N/A'],
                ['Alternative Names (SAN)', params.san or 'N/A'],
                ['Issuer', params.issuer or 'N/A'],
                ['Subject', params.subject or 'N/A'],
                ['Serial Number', params.serial_number or 'N/A'],
                ['Versión', params.version or 'N/A'],
                ['Algoritmo de Firma', params.algoritmo_firma or 'N/A'],
                ['Public Key', f'{params.key_algorithm} {params.key_size} bits' if params.key_algorithm else 'N/A'],
                ['Fecha Inicio', params.fecha_inicio.strftime('%d/%m/%Y') if params.fecha_inicio else 'N/A'],
                ['Fecha Expiración', params.fecha_fin.strftime('%d/%m/%Y') if params.fecha_fin else 'N/A'],
                ['Días Restantes', str(params.dias_restantes) if params.dias_restantes is not None else 'N/A'],
                ['Estado Revocación', params.estado_revocacion or 'N/A'],
            ]
            
            story.append(Paragraph('A) Parámetros Generales del Certificado', styles['Heading2']))
            general_table = Table(general_data, colWidths=[2.5*inch, 3.5*inch])
            general_table.setStyle(self._get_table_style())
            story.append(general_table)
            story.append(Spacer(1, 12))
        
        # B) Evaluación de Protocolos y Cifradores SSL/TLS
        if hasattr(analysis, 'parametros_tls') and analysis.parametros_tls:
            tls_params = analysis.parametros_tls
            
            protocols_supported = []
            if tls_params.tls13_supported: protocols_supported.append('TLS 1.3')
            if tls_params.tls12_supported: protocols_supported.append('TLS 1.2')
            if tls_params.tls11_supported: protocols_supported.append('TLS 1.1')
            if tls_params.tls10_supported: protocols_supported.append('TLS 1.0')
            if tls_params.sslv3_supported: protocols_supported.append('SSLv3')
            if tls_params.sslv2_supported: protocols_supported.append('SSLv2')
            
            tls_data = [
                ['Parámetro', 'Valor'],
                ['Protocolos Soportados', ', '.join(protocols_supported) if protocols_supported else 'N/A'],
                ['Cifrados Disponibles', tls_params.cifrados_disponibles[:100] + '...' if len(tls_params.cifrados_disponibles or '') > 100 else (tls_params.cifrados_disponibles or 'N/A')],
                ['Cifrados Débiles', tls_params.cifrados_debiles or 'Ninguno detectado'],
                ['Perfect Forward Secrecy', 'Sí' if tls_params.pfs else 'No'],
            ]
            
            story.append(Paragraph('B) Evaluación de Protocolos y Cifradores SSL/TLS', styles['Heading2']))
            tls_table = Table(tls_data, colWidths=[2.5*inch, 3.5*inch])
            tls_table.setStyle(self._get_table_style())
            story.append(tls_table)
            story.append(Spacer(1, 12))
        
        # C) Vulnerabilidades Conocidas
        vulnerabilities = analysis.vulnerabilidades.all()
        if vulnerabilities.exists():
            vuln_data = [['Vulnerabilidad', 'Severidad', 'Descripción']]
            
            # Mapear vulnerabilidades según especificaciones
            vuln_mapping = {
                'HEARTBLEED': 'Vulnerabilidad crítica en OpenSSL (CVE-2014-0160)',
                'POODLE': 'Vulnerabilidad asociada al soporte SSLv3 (CVE-2014-3566)',
                'DROWN': 'Ataque que explota SSLv2 (CVE-2016-0800)',
                'BEAST': 'Vulnerabilidades conocidas de ataques contra TLS',
                'CRIME': 'Vulnerabilidades conocidas de ataques contra TLS',
                'BREACH': 'Vulnerabilidades conocidas de ataques contra TLS',
                'LOGJAM': 'Debilidades en cifrados por intercambio de claves débiles',
                'FREAK': 'Debilidades en cifrados por intercambio de claves débiles',
            }
            
            for vuln in vulnerabilities:
                description = vuln_mapping.get(vuln.vulnerabilidad, vuln.description)
                vuln_data.append([
                    vuln.vulnerabilidad,
                    vuln.severity,
                    description[:60] + '...' if len(description) > 60 else description
                ])
            
            story.append(Paragraph('C) Vulnerabilidades Conocidas', styles['Heading2']))
            vuln_table = Table(vuln_data, colWidths=[1.5*inch, 1*inch, 3.5*inch])
            vuln_table.setStyle(self._get_vulnerability_table_style())
            story.append(vuln_table)
            story.append(Spacer(1, 12))
        
        # D) Validación e Integridad de la Cadena de Certificación
        if hasattr(analysis, 'cadena_certificacion') and analysis.cadena_certificacion:
            chain = analysis.cadena_certificacion
            
            chain_data = [
                ['Parámetro', 'Estado'],
                ['Cadena de confianza completa', 'Válida' if chain.cadena_ok else 'Inválida'],
                ['Errores de Cadena', 'Sí' if chain.errores else 'No'],
                ['Auto-firmado', 'Sí' if chain.autofirmado else 'No'],
                ['Intermedios adicionales', 'Sí' if chain.intermedios_extra else 'No'],
            ]
            
            story.append(Paragraph('D) Validación e Integridad de la Cadena de Certificación', styles['Heading2']))
            chain_table = Table(chain_data, colWidths=[3*inch, 2*inch])
            chain_table.setStyle(self._get_table_style())
            story.append(chain_table)
            story.append(Spacer(1, 12))
        
        # F) Parámetros adicionales relevantes para APIs y Web
        if hasattr(analysis, 'parametros_web') and analysis.parametros_web:
            web_params = analysis.parametros_web
            
            web_data = [
                ['Parámetro', 'Estado'],
                ['HSTS (Strict-Transport-Security)', 'Habilitado' if web_params.hsts else 'Deshabilitado'],
                ['Expect-CT', 'Habilitado' if web_params.expect_ct else 'Deshabilitado'],
                ['HPKP (Public-Key-Pins)', 'Habilitado' if web_params.hpkp else 'Deshabilitado'],
                ['SNI (Server Name Indication)', 'Soportado' if web_params.sni else 'No soportado'],
                ['OCSP Stapling', 'Habilitado' if web_params.ocsp_stapling else 'Deshabilitado'],
            ]
            
            story.append(Paragraph('F) Parámetros adicionales relevantes para APIs y Web', styles['Heading2']))
            web_table = Table(web_data, colWidths=[3*inch, 2*inch])
            web_table.setStyle(self._get_table_style())
            story.append(web_table)
            story.append(Spacer(1, 12))
        
        # G) Otros parámetros Útiles
        if hasattr(analysis, 'otros_parametros') and analysis.otros_parametros:
            other_params = analysis.otros_parametros
            
            other_data = [
                ['Parámetro', 'Valor'],
                ['Tiempo de respuesta SSL', f'{other_params.tiempo_respuesta_ssl} ms' if other_params.tiempo_respuesta_ssl else 'N/A'],
                ['Disponibilidad del servidor', 'Activo' if other_params.disponibilidad else 'Inactivo'],
                ['Tiempo Handshake', f'{other_params.handshake_time_ms} ms' if other_params.handshake_time_ms else 'N/A'],
            ]
            
            story.append(Paragraph('G) Otros parámetros Útiles', styles['Heading2']))
            other_table = Table(other_data, colWidths=[3*inch, 2*inch])
            other_table.setStyle(self._get_table_style())
            story.append(other_table)
            story.append(Spacer(1, 12))
        
        return story
    
    def _get_table_style(self):
        """Estilo estándar para tablas"""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
    
    def _get_vulnerability_table_style(self):
        """Estilo para tabla de vulnerabilidades"""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.red),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightcoral),
        ])
    
    def _get_expiring_certificates_count(self, days):
        """Contar certificados que expiran en X días"""
        from datetime import date, timedelta
        cutoff_date = date.today() + timedelta(days=days)
        
        return self.get_certificates_queryset().filter(
            analyses__parametros_generales__fecha_fin__lte=cutoff_date,
            analyses__tuvo_exito=True
        ).distinct().count()
    
    def generate_client_summary_report(self, file_path):
        """Reporte resumido por cliente según especificaciones"""
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Título
        story.append(Paragraph(f'Reporte Resumido - {self.client.name}', styles['Title']))
        story.append(Paragraph(f'Generado: {self.created_at.strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Resumen de certificados con información visual
        certificates = self.get_certificates_queryset()
        
        cert_data = [['Cliente', 'IP/URL', 'Puerto', 'Estado Vitalidad', 'Última Revisión', 'Próxima Revisión', 'Expiración']]
        
        for cert in certificates:
            last_analysis = cert.analyses.filter(tuvo_exito=True).first()
            
            # Estado de vitalidad (verde/rojo)
            vitality_status = 'Verde' if (last_analysis and 
                                       hasattr(last_analysis, 'otros_parametros') and 
                                       last_analysis.otros_parametros and 
                                       last_analysis.otros_parametros.disponibilidad) else 'Rojo'
            
            # Fechas según especificaciones
            last_review = last_analysis.fecha_inicio.strftime('%d/%m/%Y') if last_analysis else 'N/A'
            next_review = 'Próxima programada'  # Calcular según frecuencia
            
            expiration = 'N/A'
            if (last_analysis and hasattr(last_analysis, 'parametros_generales') and 
                last_analysis.parametros_generales and last_analysis.parametros_generales.fecha_fin):
                exp_date = last_analysis.parametros_generales.fecha_fin
                days_remaining = (exp_date - date.today()).days if exp_date else None
                expiration = f'{exp_date.strftime("%d/%m/%Y")} ({days_remaining} días)' if days_remaining is not None else exp_date.strftime('%d/%m/%Y')
            
            cert_data.append([
                self.client.name,
                cert.ip or cert.url,
                str(cert.puerto),
                vitality_status,
                last_review,
                next_review,
                expiration
            ])
        
        cert_table = Table(cert_data, colWidths=[1*inch, 1.5*inch, 0.7*inch, 1*inch, 1*inch, 1*inch, 1.8*inch])
        cert_table.setStyle(self._get_table_style())
        
        story.append(Paragraph('Estado de Certificados', styles['Heading2']))
        story.append(cert_table)
        
        doc.build(story)
        return file_path


class ExcelReportGenerator(ReportGenerator):
    """
    Generador de reportes Excel según especificaciones Proyecto Sócrates
    """
    
    def generate_client_detailed_excel(self, file_path):
        """Generar Excel detallado con todas las especificaciones"""
        wb = openpyxl.Workbook()
        
        # Hoja 1: Resumen
        ws_summary = wb.active
        ws_summary.title = 'Resumen'
        self._create_summary_sheet(ws_summary)
        
        # Hoja 2: Parámetros Generales
        ws_general = wb.create_sheet('Parámetros Generales')
        self._create_general_params_sheet(ws_general)
        
        # Hoja 3: Protocolos SSL/TLS
        ws_tls = wb.create_sheet('SSL/TLS')
        self._create_tls_sheet(ws_tls)
        
        # Hoja 4: Vulnerabilidades
        ws_vulns = wb.create_sheet('Vulnerabilidades')
        self._create_vulnerabilities_sheet(ws_vulns)
        
        # Hoja 5: Estado de Certificados
        ws_status = wb.create_sheet('Estado Certificados')
        self._create_status_sheet(ws_status)
        
        wb.save(file_path)
        return file_path
    
    def _create_summary_sheet(self, ws):
        """Crear hoja de resumen ejecutivo"""
        ws['A1'] = f'Proyecto Sócrates - Reporte Cliente: {self.client.name}'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f'Generado: {self.created_at.strftime("%d/%m/%Y %H:%M")}'
        
        certificates = self.get_certificates_queryset()
        analyses = self.get_analyses_queryset()
        vulnerabilities = Vulnerabilidades.objects.filter(analisis__in=analyses)
        
        # Métricas principales
        metrics = [
            ('Total Certificados', certificates.count()),
            ('Análisis Realizados', analyses.count()),
            ('Vulnerabilidades Críticas', vulnerabilities.filter(severity='CRITICAL').count()),
            ('Vulnerabilidades Altas', vulnerabilities.filter(severity='HIGH').count()),
            ('Certificados Activos', certificates.filter(analyses__otros_parametros__disponibilidad=True).distinct().count()),
        ]
        
        row = 4
        ws[f'A{row}'] = 'Métrica'
        ws[f'B{row}'] = 'Valor'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        for metric, value in metrics:
            row += 1
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
    
    def _create_general_params_sheet(self, ws):
        """Crear hoja con parámetros generales según especificaciones"""
        headers = [
            'Certificado', 'Common Name (CN)', 'Alternative Names (SAN)', 'Issuer', 
            'Subject', 'Serial Number', 'Versión', 'Algoritmo Firma', 'Public Key',
            'Fecha Inicio', 'Fecha Expiración', 'Días Restantes', 'Estado Revocación'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        analyses = self.get_analyses_queryset()
        row = 2
        
        for analysis in analyses:
            if hasattr(analysis, 'parametros_generales') and analysis.parametros_generales:
                params = analysis.parametros_generales
                cert_target = analysis.certificado.ip or analysis.certificado.url
                
                values = [
                    f'{cert_target}:{analysis.certificado.puerto}',
                    params.common_name or 'N/A',
                    params.san or 'N/A',
                    params.issuer or 'N/A',
                    params.subject or 'N/A',
                    params.serial_number or 'N/A',
                    params.version or 'N/A',
                    params.algoritmo_firma or 'N/A',
                    f'{params.key_algorithm} {params.key_size}' if params.key_algorithm else 'N/A',
                    params.fecha_inicio.strftime('%d/%m/%Y') if params.fecha_inicio else 'N/A',
                    params.fecha_fin.strftime('%d/%m/%Y') if params.fecha_fin else 'N/A',
                    str(params.dias_restantes) if params.dias_restantes is not None else 'N/A',
                    params.estado_revocacion or 'N/A'
                ]
                
                for col, value in enumerate(values, 1):
                    ws.cell(row=row, column=col, value=value)
                
                row += 1
    
    def _create_tls_sheet(self, ws):
        """Crear hoja con evaluación de protocolos SSL/TLS"""
        headers = [
            'Certificado', 'TLS 1.3', 'TLS 1.2', 'TLS 1.1', 'TLS 1.0', 'SSLv3', 'SSLv2',
            'Perfect Forward Secrecy', 'Cifrados Disponibles', 'Cifrados Débiles'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        
        analyses = self.get_analyses_queryset()
        row = 2
        
        for analysis in analyses:
            if hasattr(analysis, 'parametros_tls') and analysis.parametros_tls:
                tls = analysis.parametros_tls
                cert_target = analysis.certificado.ip or analysis.certificado.url
                
                values = [
                    f'{cert_target}:{analysis.certificado.puerto}',
                    'Sí' if tls.tls13_supported else 'No',
                    'Sí' if tls.tls12_supported else 'No',
                    'Sí' if tls.tls11_supported else 'No',
                    'Sí' if tls.tls10_supported else 'No',
                    'Sí' if tls.sslv3_supported else 'No',
                    'Sí' if tls.sslv2_supported else 'No',
                    'Sí' if tls.pfs else 'No',
                    (tls.cifrados_disponibles or 'N/A')[:100],  # Truncar para Excel
                    tls.cifrados_debiles or 'Ninguno'
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    
                    # Colorear protocolos inseguros
                    if col in [6, 7] and value == 'Sí':  # SSLv3, SSLv2
                        cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                
                row += 1
    
    def _create_vulnerabilities_sheet(self, ws):
        """Crear hoja con vulnerabilidades conocidas"""
        headers = ['Certificado', 'Vulnerabilidad', 'Severidad', 'Descripción', 'Fecha Detección']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        
        analyses = self.get_analyses_queryset()
        row = 2
        
        for analysis in analyses:
            vulnerabilities = analysis.vulnerabilidades.all()
            for vuln in vulnerabilities:
                cert_target = analysis.certificado.ip or analysis.certificado.url
                
                values = [
                    f'{cert_target}:{analysis.certificado.puerto}',
                    vuln.vulnerabilidad,
                    vuln.severity,
                    vuln.description,
                    analysis.fecha_inicio.strftime('%d/%m/%Y')
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    
                    # Color según severidad
                    if vuln.severity == 'CRITICAL':
                        fill_color = 'FF9999'
                    elif vuln.severity == 'HIGH':
                        fill_color = 'FFCCCC'
                    else:
                        fill_color = 'FFFFFF'
                    
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                row += 1
    
    def _create_status_sheet(self, ws):
        """Crear hoja con estado de certificados según especificaciones"""
        headers = [
            'Cliente', 'IP/URL', 'Puerto', 'Estado Vitalidad', 'Fecha Última Revisión', 
            'Fecha Próxima Revisión', 'Fecha Expiración', 'Días para Expirar'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        
        certificates = self.get_certificates_queryset()
        row = 2
        
        for cert in certificates:
            last_analysis = cert.analyses.filter(tuvo_exito=True).first()
            
            # Estado de vitalidad según especificaciones
            vitality = 'Verde'
            if (last_analysis and hasattr(last_analysis, 'otros_parametros') and 
                last_analysis.otros_parametros and not last_analysis.otros_parametros.disponibilidad):
                vitality = 'Rojo'
            
            # Fechas
            last_review = last_analysis.fecha_inicio.strftime('%d/%m/%Y') if last_analysis else 'N/A'
            next_review = 'Programada'  # Implementar lógica de frecuencia
            
            expiration_date = ''
            days_remaining = ''
            if (last_analysis and hasattr(last_analysis, 'parametros_generales') and 
                last_analysis.parametros_generales and last_analysis.parametros_generales.fecha_fin):
                exp_date = last_analysis.parametros_generales.fecha_fin
                expiration_date = exp_date.strftime('%d/%m/%Y')
                days_remaining = last_analysis.parametros_generales.dias_restantes or 0
            
            values = [
                self.client.name,
                cert.ip or cert.url,
                str(cert.puerto),
                vitality,
                last_review,
                next_review,
                expiration_date,
                str(days_remaining) if days_remaining != '' else 'N/A'
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                
                # Destacar certificados próximos a expirar
                if col == 8 and days_remaining != '' and isinstance(days_remaining, int) and days_remaining <= 30:
                    cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            
            row += 1
